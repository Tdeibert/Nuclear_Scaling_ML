#!/usr/bin/env python3
"""Import Nuclear_Scaling_Database_Schema.xlsx data into SQLite."""

from __future__ import annotations

import argparse
import ast
import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required for Excel import. Install it with: "
        "python -m pip install openpyxl"
    ) from exc

from create_database import DEFAULT_DATABASE, create_database


SHEET_ORDER = (
    "Experimental_Cfg",
    "Nuclei",
    "NucleusZStack",
    "RawIntensities",
    "RadialProfile",
)

TABLE_CONFIG = {
    "Experimental_Cfg": {
        "table": "experimental_cfg",
        "conflict": ("experiment_id", "fov_id", "droplet_id"),
        "mapping": {
            "Experiment_ID": "experiment_id",
            "FOV_ID": "fov_id",
            "Droplet_ID": "droplet_id",
            "Date": "experiment_date",
            "Pixel_Size_um": "pixel_size_um",
            "Z_Step_um": "z_step_um",
            "Num_Z_Planes": "num_z_planes",
            "Frame_Interval_min": "frame_interval_min",
            "Channel0_Label": "channel0_label",
            "Channel1_Label": "channel1_label",
            "Channel2_Label": "channel2_label",
            "Microscope": "microscope",
            "Objective": "objective",
            "Operator": "operator",
            "Segmentation_Pipeline_Version": "segmentation_pipeline_version",
            "Model_Version": "model_version",
            "Raw_File_Path": "raw_file_path",
            "Treatment": "treatment",
            "Concentration": "concentration",
            "Concentration_Units": "concentration_units",
            "Biological_Replicate": "biological_replicate",
            "Technical_Replicate": "technical_replicate",
            "Extract_Batch": "extract_batch",
            "Experimental_Group": "experimental_group",
            "Notes": "notes",
        },
    },
    "Nuclei": {
        "table": "nuclei",
        "conflict": ("nucleus_id", "time_frame"),
        "mapping": {
            "Nucleus_ID": "nucleus_id",
            "Droplet_ID": "droplet_id",
            "FOV_ID": "fov_id",
            "Experiment_ID": "experiment_id",
            "Time_Frame": "time_frame",
            "Time_Real_Minutes": "time_real_minutes",
            "Stage_Classification": "stage_classification",
            "Selected_Slice_ID": "selected_slice_id",
            "Cross_Sectional_Area_um2": "cross_sectional_area_um2",
            "Volume_3D_um3": "volume_3d_um3",
            "NC_Ratio": "nc_ratio",
            "Segmentation_Pipeline_Version": "segmentation_pipeline_version",
            "Model_Version": "model_version",
            "QC_Flag": "qc_flag",
            "QC_Notes": "qc_notes",
            "Source_File_Path": "source_file_path",
        },
        "centroid": "Centroid_XYZ_px",
    },
    "NucleusZStack": {
        "table": "nucleus_z_stack",
        "conflict": ("nucleus_id", "time_frame", "slice_id"),
        "mapping": {
            "Nucleus_ID": "nucleus_id",
            "Time_Frame": "time_frame",
            "Slice_ID": "slice_id",
            "Cross_Sectional_Area_um2": "cross_sectional_area_um2",
            "Is_Selected_Max": "is_selected_max",
        },
        "centroid": "Centroid_XYZ_px",
    },
    "RawIntensities": {
        "table": "raw_intensities",
        "conflict": ("nucleus_id", "time_frame"),
        "mapping": {
            "Nucleus_ID": "nucleus_id",
            "Time_Frame": "time_frame",
            **{
                f"Halo{halo}_{channel}": f"halo{halo}_{channel.lower()}"
                for halo in range(1, 5)
                for channel in ("Mcherry", "NPC", "Membrane")
            },
            "Background_Mcherry": "background_mcherry",
            "Background_NPC": "background_npc",
            "Background_Membrane": "background_membrane",
        },
    },
    "RadialProfile": {
        "table": "radial_profile",
        "conflict": (
            "nucleus_id",
            "time_frame",
            "theta_deg",
            "rho_normalized",
            "channel",
        ),
        "mapping": {
            "Nucleus_ID": "nucleus_id",
            "Time_Frame": "time_frame",
            "Theta_deg": "theta_deg",
            "Rho_Normalized": "rho_normalized",
            "W_Wall_Proximity_Index": "w_wall_proximity_index",
            "Channel": "channel",
            "Intensity": "intensity",
            "Is_Ridge_Point": "is_ridge_point",
            "Cluster_ID": "cluster_id",
        },
    },
}

BOOLEAN_COLUMNS = {"is_selected_max", "is_ridge_point"}


def clean_value(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value.strip() if isinstance(value, str) else value


def parse_boolean(value: Any, field: str) -> int:
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)) and value in (0, 1):
        return int(value)
    normalized = str(value).strip().lower()
    if normalized in {"true", "t", "yes", "y", "1"}:
        return 1
    if normalized in {"false", "f", "no", "n", "0"}:
        return 0
    raise ValueError(f"Invalid Boolean value for {field}: {value!r}")


def parse_centroid(value: Any) -> tuple[float | None, float | None, float | None]:
    if value is None or value == "":
        return None, None, None
    parsed = ast.literal_eval(str(value))
    if not isinstance(parsed, (tuple, list)) or len(parsed) != 3:
        raise ValueError(f"Centroid must contain exactly three values: {value!r}")
    return tuple(float(component) for component in parsed)


def worksheet_records(worksheet: Any) -> Iterable[dict[str, Any]]:
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration:
        return
    if not any(headers):
        return
    for row_number, row in enumerate(rows, start=2):
        if not any(value is not None and value != "" for value in row):
            continue
        yield {"__row__": row_number, **dict(zip(headers, row))}


def transform_record(sheet: str, source: dict[str, Any]) -> dict[str, Any]:
    config = TABLE_CONFIG[sheet]
    record = {
        target: clean_value(source.get(origin))
        for origin, target in config["mapping"].items()
        if origin in source
    }
    for column in BOOLEAN_COLUMNS.intersection(record):
        record[column] = parse_boolean(record[column], column)
    if centroid_column := config.get("centroid"):
        x, y, z = parse_centroid(source.get(centroid_column))
        record.update(centroid_x_px=x, centroid_y_px=y, centroid_z_px=z)
    return record


def upsert(connection: sqlite3.Connection, table: str, conflict: tuple[str, ...], record: dict[str, Any]) -> None:
    columns = tuple(record)
    updates = tuple(column for column in columns if column not in conflict)
    placeholders = ", ".join("?" for _ in columns)
    conflict_sql = ", ".join(conflict)
    if updates:
        action = "DO UPDATE SET " + ", ".join(
            f"{column} = excluded.{column}" for column in updates
        )
    else:
        action = "DO NOTHING"
    sql = (
        f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders}) "
        f"ON CONFLICT ({conflict_sql}) {action}"
    )
    connection.execute(sql, tuple(record[column] for column in columns))


def import_workbook(workbook_path: Path, database_path: Path) -> dict[str, int]:
    workbook_path = workbook_path.expanduser().resolve()
    database_path = database_path.expanduser().resolve()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    create_database(database_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    missing = [sheet for sheet in SHEET_ORDER if sheet not in workbook.sheetnames]
    if missing:
        raise ValueError(f"Workbook is missing required sheets: {', '.join(missing)}")

    counts: dict[str, int] = {}
    with sqlite3.connect(database_path) as connection:
        connection.execute("PRAGMA foreign_keys = ON")
        try:
            for sheet in SHEET_ORDER:
                config = TABLE_CONFIG[sheet]
                count = 0
                for source in worksheet_records(workbook[sheet]):
                    try:
                        record = transform_record(sheet, source)
                        upsert(connection, config["table"], config["conflict"], record)
                    except Exception as exc:
                        raise ValueError(f"{sheet} row {source['__row__']}: {exc}") from exc
                    count += 1
                counts[sheet] = count

            violations = connection.execute("PRAGMA foreign_key_check").fetchall()
            if violations:
                raise RuntimeError(f"Foreign-key violations detected: {violations[:10]}")
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            workbook.close()
    return counts


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Path to the XLSX workbook")
    parser.add_argument(
        "--database",
        type=Path,
        default=DEFAULT_DATABASE,
        help=f"Database path (default: {DEFAULT_DATABASE})",
    )
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_args()
    imported = import_workbook(arguments.workbook, arguments.database)
    print("Import completed:")
    for sheet_name, row_count in imported.items():
        print(f"  {sheet_name}: {row_count} row(s)")
